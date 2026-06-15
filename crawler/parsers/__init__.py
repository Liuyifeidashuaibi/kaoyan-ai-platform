"""多格式解析器。"""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path

log = logging.getLogger("crawler.parsers")


def parse_html_to_text(html: str) -> str:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    return soup.get_text("\n", strip=True)


def parse_pdf_bytes(data: bytes) -> str:
    try:
        import pdfplumber
    except ImportError:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=data, filetype="pdf")
            return "\n".join(page.get_text() for page in doc)
        except ImportError as exc:
            raise RuntimeError("请安装 pdfplumber 或 PyMuPDF") from exc

    parts: list[str] = []
    with pdfplumber.open(BytesIO(data)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def parse_docx_bytes(data: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_excel_bytes(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines)


def parse_file(path: Path) -> str:
    suffix = path.suffix.lower()
    data = path.read_bytes()
    if suffix == ".pdf":
        return parse_pdf_bytes(data)
    if suffix in (".docx", ".doc"):
        return parse_docx_bytes(data)
    if suffix in (".xlsx", ".xls"):
        return parse_excel_bytes(data)
    return data.decode("utf-8", errors="ignore")
